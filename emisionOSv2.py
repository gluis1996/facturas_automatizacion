import os
from dataclasses import dataclass
from datetime import datetime, timedelta
import PyPDF2
import pandas as pd
import time
from pdfminer.high_level import extract_text
import re
import openpyxl 
import win32com.client
import subprocess

import sys

original_sys = sys.stdout

@dataclass
class claseFactura:
    sociedad : str
    asegurado : str
    grupoPersonal : str
    detalle : str
    prima : str
    igv : str
    primaTotal : str

hojaDatos = openpyxl.load_workbook('D:\\plunarejoc\\Automatizaciones\\Emisión de Ordenes de Servicio\\90 Excel Config.xlsx')
workSheet = hojaDatos.active

listaFacturas = []
fechaActual = datetime.now()
fechaActualTexto = fechaActual.strftime('%d.%m.%Y')
fechaSiguiente = fechaActual + timedelta(7)
fechaSiguienteTexto = fechaSiguiente.strftime('%d.%m.%Y')

user = workSheet.cell(row = 6, column = 6).value
password = workSheet.cell(row = 7, column = 6).value
rutaRepositorio = workSheet.cell(row = 8, column = 6).value
rutaRepositorioMapfre = workSheet.cell(row = 9, column = 6).value
rutaRepositorioOS = workSheet.cell(row = 10, column = 6).value

hojaDatos.close()



dataFrameNuevosNombres = None
dataFrameSociedades = pd.read_excel(rutaRepositorioOS + 'Sociedades.xlsx')
print(dataFrameSociedades)

def leerNuevosNombres():
    dataFrameAuxiliar = pd.read_excel(rutaRepositorioOS + 'Nuevos nombres.xlsx')
    for index, fila in dataFrameAuxiliar.iterrows():
        nuevoNombre = fila['DETALLE'].split(' ')
        nuevoNombre[len(nuevoNombre) - 1] = datetime.now().strftime('%m.%y')
        nuevoDetalle = ''
        for elemento in nuevoNombre:
            nuevoDetalle = nuevoDetalle + ' ' + elemento + ' '
        nuevoDetalle = nuevoDetalle.replace('  ', ' ')
        dataFrameAuxiliar.loc[index, 'DETALLE'] = nuevoDetalle
    return dataFrameAuxiliar

def leerCarpetas():
    print(rutaRepositorio)
    listaCarpetas = os.listdir(rutaRepositorio)
    print('\nComenzar a revisar información de facturas PDF en carpeta')
    for carpeta in listaCarpetas:
        if carpeta.endswith('.xlsx') == False and carpeta.endswith('.csv') == False:
            listaSubCarpetas = os.listdir(rutaRepositorio + '\\'+ carpeta)
            for archivo in listaSubCarpetas:
                if archivo.endswith('.pdf'):
                    nombreSociedad = archivo.split('-')[0]
                    archivoPDF = open(rutaRepositorio + '\\'+ carpeta + '\\' + archivo, 'rb')
                    lectorPDF = PyPDF2.PdfReader(archivoPDF)
                    cantidadPaginas = len(lectorPDF.pages)
                    print('Cantidad: ', cantidadPaginas, archivo, '===================================')
                    try:
                        for paginaNumero in range(cantidadPaginas):
                            pagina = lectorPDF.pages[paginaNumero]
                            texto = pagina.extract_text()
                            texto = texto.split('\n')
                            for indice in range(len(texto)):
                                print(indice, texto[indice])
                            if texto[14].startswith('Asesor: ') == False:
                                prima = texto[33].replace(' ', '')
                                igv = texto[36].replace(' ', '')
                                primaTotal = texto[29].replace(' ', '')
                        
                            else:
                                prima = texto[34].replace(' ', '')
                                igv = texto[37].replace(' ', '')
                                primaTotal = texto[30].replace(' ', '')
                            prima = float(prima.replace(',', ''))
                            igv = float(igv.replace(',', ''))
                            primaTotal = float(primaTotal.replace(',', ''))

                            nombreArchivo = archivo.replace('.pdf', '')
                    except Exception as e:
                        print('Error encontrado: ', e)
                        time.sleep(8000)
                    listaFacturas.append(claseFactura(nombreSociedad, '', '', nombreArchivo, prima, igv, primaTotal))
    dataFrameAuxiliar = pd.DataFrame(
        [(factura.sociedad, factura.detalle, factura.prima, factura.igv, factura.primaTotal) for factura in listaFacturas],
        columns=['Sociedad', 'nombreArchivo', 'prima', 'igv', 'primaTotal']
    )
    print(dataFrameAuxiliar)
    return dataFrameAuxiliar

def leerCarpetaMAPFRE():
    print('Comenzar a leer información de facturas de MAPFRE:\n')
    listaFacturas = os.listdir(rutaRepositorioMapfre)
    listaFacturasLeidas = []
    for archivo in listaFacturas:
        if archivo.endswith('.PDF') or archivo.endswith('.pdf'):
            archivoPDF = open(rutaRepositorioMapfre + archivo, 'rb')
            lectorPDF = PyPDF2.PdfReader(archivoPDF)
            cantidadPaginas = len(lectorPDF.pages)
            try:
                texto = extract_text(rutaRepositorioMapfre + archivo)
                texto = texto.split('\n')
                for paginaNumero in range(cantidadPaginas):
                    pagina = lectorPDF.pages[paginaNumero]
                    texto = pagina.extract_text()
                    texto = texto.split('\n')
                    igv = 0
                    prima = 0
                    primaTotal = 0
                    for numeroElemento in range(len(texto)):
                        if texto[numeroElemento].startswith('Prima Comercial + IGV :'):
                            linea = texto[numeroElemento].split(' ')
                            primaTotal = float(linea[6].replace(',', ''))
                            # print('Prima total: ', primaTotal, ' archivo: ', archivo)
                            prima = primaTotal / 1.18
                            igv = prima * 0.18
                    if texto[0].startswith('VIDA'):
                        if texto[38].startswith('EMPLEADOS'):
                            detalle = 'VIDA LEY EMPLEADOS'
                            if texto[28].startswith('ALICORP') == False:
                                grupoPersonal = 'CESADO'
                            else:
                                grupoPersonal = 'EMPLEADO'
                        else:
                            detalle = 'VIDA LEY OBREROS'
                        contratante = texto[28]
                    else:
                        for numeroElemento in range(len(texto)):
                            indiceAlterno = 0
                            if texto[numeroElemento].startswith('GRUPO MAPFRE PERUEl presente documento no'):
                                indiceAlterno = numeroElemento - 2
                                detalle = texto[indiceAlterno].replace('/', '')
                                detalle = re.sub(r"\d+", "", detalle)
                                detalle = detalle.strip()
                                detalle = detalle.replace('SEGURO COMPLEMENTARIO DE TRABAJO DE RIESGO', 'SCTR')
                                detalle = detalle.replace('PENSIONES', 'PENSIÓN')

                        if texto[27].startswith('ALICORP') == False:
                            grupoPersonal = 'CESADO'
                        else:
                            grupoPersonal = 'EMPLEADO'
                        contratante = texto[27]
                    detalle = detalle + ' ' + datetime.now().strftime('%m.%y')
                    # Revisar sociedad en dataframe de facturas
                    for index, fila in dataFrameSociedades.iterrows():
                        if contratante.upper().startswith(fila['Clave'].upper()):
                            sociedad = fila['CÓDIGO']
                        
                listaFacturasLeidas.append(claseFactura(sociedad, grupoPersonal, '', detalle, prima, igv, primaTotal))
            except Exception as e:
                print('Error encontrado: ', str(e))
    dataFrameAuxiliar = pd.DataFrame(
        [(factura.sociedad, factura.detalle, factura.prima, factura.igv, factura.primaTotal) for factura in listaFacturasLeidas],
        columns=['Sociedad', 'nombreArchivo', 'prima', 'igv', 'primaTotal']
    )
    print(dataFrameAuxiliar)
    return dataFrameAuxiliar       

def leerONCOCENTER():
    print('Comenzar a leer facturas de oncocenter')
    listaArchivos =  os.listdir(rutaRepositorioOS + 'ONCOCENTER\\')
    listaFacturas = []
    for archivo in listaArchivos:
        if archivo.upper().endswith('PDF'):
            texto = extract_text(rutaRepositorioOS + 'ONCOCENTER\\' + archivo, 'rb')
            texto = texto.split('\n')
            try:
                prima = float(texto[135].replace(',',''))
                igv = prima * 0.18
                primaTotal = prima * 1.18
                listaFacturas.append(claseFactura('PE11', '', 'EMPLEADO', 'Chequeo médico ' + datetime.now().strftime('%m.%y'), prima, igv, primaTotal))
            except Exception as e:
                print('Error encontrado: ', e)
    dataFrameAuxiliar = pd.DataFrame(
        [(factura.sociedad, factura.detalle, factura.prima, factura.igv, factura.primaTotal) for factura in listaFacturas],
        columns=['Sociedad', 'nombreArchivo', 'prima', 'igv', 'primaTotal']
    ) 
    return dataFrameAuxiliar

def leerTEBCA():
    print('Comenzar a leer facturas de TEBCA:')
    listaArchivos =  os.listdir(rutaRepositorioOS + 'TEBCA\\COMISION\\')
    listaFacturas = []
    for archivo in listaArchivos:
        if archivo.upper().endswith('PDF'):
            texto = extract_text(rutaRepositorioOS + 'TEBCA\\COMISION\\' + archivo, 'rb')
            texto = texto.split('\n')
            try:
                prima = float(texto[88].replace(',','').replace('S/', ''))
                igv = prima * 0.18
                primaTotal = prima * 1.18
                listaFacturas.append(claseFactura('PE14', '', 'EMPLEADO', 'RECARGA VALES COMISION ' + datetime.now().strftime('%m.%y'), prima, igv, primaTotal))
            except Exception as e:
                print('Error encontrado: ', e)

    listaArchivos =  os.listdir(rutaRepositorioOS + 'TEBCA\\RECARGA VALES VISA\\')
    for archivo in listaArchivos:
        if archivo.upper().endswith('PDF'):
            texto = extract_text(rutaRepositorioOS + 'TEBCA\\RECARGA VALES VISA\\' + archivo, 'rb')
            texto = texto.split('\n')
            try:
                prima = float(texto[97].replace(',','').replace('S/', ''))
                igv = prima * 0.18
                primaTotal = prima * 1.18
                listaFacturas.append(claseFactura('PE14', '', 'EMPLEADO', 'RECARGA VALES VISA ' + datetime.now().strftime('%m.%y'), prima, igv, primaTotal))
            except Exception as e:
                print('Error encontrado: ', e)
    dataFrameAuxiliar = pd.DataFrame(
        [(factura.sociedad, factura.detalle, factura.prima, factura.igv, factura.primaTotal) for factura in listaFacturas],
        columns=['Sociedad', 'nombreArchivo', 'prima', 'igv', 'primaTotal']
    )
    return dataFrameAuxiliar
    
def adicionarInformacion(dataFrameFacturas, dataFrameFacturasMAPFRE, dataFrameOncocenter, dataFrameTEBCA):
    print('\nComenzar a modificar información de dataFrame')
    print('MAPFRE: ')
    print('Cantidad de facturas: ', len(dataFrameFacturasMAPFRE))
    for index, fila in dataFrameFacturasMAPFRE.iterrows():
        for index2, fila2 in dataFrameNuevosNombres.iterrows():
            nombreArchivoAuxiliar = str(fila['nombreArchivo']).replace(' ', '').lstrip()
            detalleAuxiliar = str(fila2['DETALLE']).replace(' ', '').lstrip()
            if nombreArchivoAuxiliar == detalleAuxiliar.lstrip():
                dataFrameFacturasMAPFRE.loc[index, 'nombreArchivo'] = fila2['DETALLE']
                dataFrameFacturasMAPFRE.loc[index, 'Imputacion'] = fila2['I']
                dataFrameFacturasMAPFRE.loc[index, 'Grupo'] = fila2['Grupo de Personal']
                dataFrameFacturasMAPFRE.loc[index, 'Anticipo'] = fila2['ANTICIPO']
                dataFrameFacturasMAPFRE.loc[index, 'Incluye'] = fila2['Incluye']
                dataFrameFacturasMAPFRE.loc[index, 'Proveedor'] = fila2['Proveedor']
                dataFrameFacturasMAPFRE.loc[index, 'Codigo Proveedor'] = fila2['Cod. Proveedor']
    print('Data adicional de facturas de MAPFRE añadida')
            
    print('PACIFICO: ')
    print('Cantidad de facturas: ', len(dataFrameFacturas))
    print(dataFrameFacturas)
    for index, fila in dataFrameFacturas.iterrows():
        nombreArchivoAuxiliar = fila['nombreArchivo']
        nombreArchivoAuxiliar = '-'.join(str(nombreArchivoAuxiliar).split('-')[1:])
        nombreArchivoAuxiliar = nombreArchivoAuxiliar.replace(' ', '')
        # print(nombreArchivoAuxiliar)
        for index2, fila2 in dataFrameNuevosNombres.iterrows():
            if fila2['Proveedor'] == 'PACIFICO':
                distintivoAuxiliar = fila2['DISTINTIVO FACT']
                # print(distintivoAuxiliar, index2)
                distintivoAuxiliar = '-'.join(str(distintivoAuxiliar).split('-')[1:])
                distintivoAuxiliar = distintivoAuxiliar.replace(' ', '')
                # print(distintivoAuxiliar, index2)
                if nombreArchivoAuxiliar == distintivoAuxiliar:
                    dataFrameFacturas.loc[index, 'nombreArchivo'] = fila2['DETALLE']
                    dataFrameFacturas.loc[index, 'Imputacion'] = fila2['I']
                    dataFrameFacturas.loc[index, 'Grupo'] = fila2['Grupo de Personal']
                    dataFrameFacturas.loc[index, 'Anticipo'] = fila2['ANTICIPO']
                    dataFrameFacturas.loc[index, 'Incluye'] = fila2['Incluye']
                    dataFrameFacturas.loc[index, 'Proveedor'] = fila2['Proveedor']
                    dataFrameFacturas.loc[index, 'Codigo Proveedor'] = fila2['Cod. Proveedor']
                # if nombreArchivoAuxiliar.startswith('EPS-'): 
                    # print('Match: ', nombreArchivoAuxiliar, ' == ', distintivoAuxiliar, index2)
        # print(nombreArchivoAuxiliar)

    # time.sleep(980)   
    # print(dataFrameFacturas)
    print('Data adicional de facturas de PACIFICO añadida')

    print('ONCOCENTER: ')
    print('Cantidad de facturas: ', len(dataFrameOncocenter))
    for index, fila in dataFrameOncocenter.iterrows():
        dataFrameOncocenter.loc[index, 'nombreArchivo'] = 'Chequeo médico ' + datetime.now().strftime('%m.%y')
        dataFrameOncocenter.loc[index, 'Imputacion'] = 'K'
        dataFrameOncocenter.loc[index, 'Grupo'] = 'EMPLEADO'
        dataFrameOncocenter.loc[index, 'Anticipo'] = 'NO'
        dataFrameOncocenter.loc[index, 'Incluye'] = 'Prima sin IGV'   
        dataFrameOncocenter.loc[index, 'Proveedor'] = 'ONCOCENTER PERÚ'
        dataFrameOncocenter.loc[index, 'Codigo Proveedor'] = 1000002873
    print('Data adicional de facturas de ONCOCENTER añadida')

    print('TEBCA: ')
    print('Cantidad de facturas: ', len(dataFrameTEBCA))    
    for index, fila in dataFrameTEBCA.iterrows():
        # dataFrameTEBCA.loc[index, 'nombreArchivo'] = 'Chequeo médico ' + datetime.now().strftime('%m.%y')
        if fila['nombreArchivo'] == 'RECARGA VALES VISA ' + datetime.now().strftime('%m.%y'):
            dataFrameTEBCA.loc[index, 'Imputacion'] = 'H'
            dataFrameTEBCA.loc[index, 'Anticipo'] = 'SI'
            dataFrameTEBCA.loc[index, 'Incluye'] = 'Total Recargas' 
        else:
            dataFrameTEBCA.loc[index, 'Imputacion'] = 'K'
            dataFrameTEBCA.loc[index, 'Anticipo'] = 'NO'
            dataFrameTEBCA.loc[index, 'Incluye'] = 'Prima sin IGV'   
            
        dataFrameTEBCA.loc[index, 'Grupo'] = 'EMPLEADO'
        dataFrameTEBCA.loc[index, 'Proveedor'] = 'TEBCA'
        dataFrameTEBCA.loc[index, 'Codigo Proveedor'] = 1000002873
    print('Data adicional de facturas de TEBCA añadida')
            

    # Cambiar nombre de columna
    dataFrameFacturas = dataFrameFacturas.rename(columns = {'nombreArchivo' : 'Detalle'})
    dataFrameFacturasMAPFRE = dataFrameFacturasMAPFRE.rename(columns = {'nombreArchivo' : 'Detalle'})
    dataFrameOncocenter = dataFrameOncocenter.rename(columns = {'nombreArchivo' : 'Detalle'})
    dataFrameTEBCA = dataFrameTEBCA.rename(columns = {'nombreArchivo' : 'Detalle'})

    dataFrameAuxiliar = pd.concat([dataFrameFacturasMAPFRE, dataFrameFacturas, dataFrameOncocenter, dataFrameTEBCA], ignore_index=True)
    dataFrameAuxiliar = dataFrameAuxiliar.rename(columns = {'prima' : 'Prima'})
    dataFrameAuxiliar = dataFrameAuxiliar.rename(columns = {'primaTotal' : 'Prima total'})
    dataFrameAuxiliar = dataFrameAuxiliar.rename(columns = {'igv' : 'IGV'})

    # print(dataFrameAuxiliar)
    # time.sleep(80)

    for index, fila in dataFrameAuxiliar.iterrows():
        dataFrameAuxiliar.loc[index, 'Prima total'] = "{:,.2f}".format(fila['Prima total'])
        dataFrameAuxiliar.loc[index, 'Prima'] = "{:,.2f}".format(fila['Prima'])
        dataFrameAuxiliar.loc[index, 'IGV'] = "{:,.2f}".format(fila['IGV'])
    print('\nDataFrame combinado: ')
    print(dataFrameAuxiliar)
    dataFrameAuxiliar.to_excel(rutaRepositorioOS + '11 Generado\\Facturas.xlsx', index=False)
    print('Se ha actualizado archivo de excel')

def leerGrupoArticulo():
    dataFrameGrupo = pd.read_excel(rutaRepositorioOS + 'Grupo articulo y varios.xlsx')
    # print(dataFrameGrupo)
    dataFrameFacturas = pd.read_excel(rutaRepositorioOS + '11 Generado\\Facturas.xlsx')

    for index, fila in dataFrameFacturas.iterrows():
        dataFrameFacturas.loc[index, 'Grupo Articulo'] = ''
        dataFrameFacturas.loc[index, 'Codigo de Grupo Articulo'] = ''
        dataFrameFacturas.loc[index, 'Centro'] = ''
        dataFrameFacturas.loc[index, 'Codigo de centro'] = ''
        dataFrameFacturas.loc[index, 'Servicio'] = ''
        dataFrameFacturas.loc[index, 'CECO'] = ''


    for index, fila in dataFrameFacturas.iterrows():
        for index2, fila2 in dataFrameGrupo.iterrows():
            # print(fila2['DETALLE'])
            if fila['Sociedad'] == fila2['Codigo de Sociedad']:
                detalleAuxiliar = str(fila['Detalle']).split()
                detalleAuxiliar = ''.join(detalleAuxiliar[:-1])

                detalleAuxuliarGrupo = str(fila2['DETALLE']).replace(' ', '')
                # detalleAuxuliarGrupo = ''.join(detalleAuxuliarGrupo[:-])
                
                # print(detalleAuxiliar, detalleAuxuliarGrupo)
                # print(detalleAuxiliar, detalleAuxuliarGrupo, index2, fila['Sociedad'], fila2['Codigo de Sociedad'])
                if detalleAuxiliar == detalleAuxuliarGrupo:
            # fila['Detalle'].startswith(fila2['DETALLE']) a:
                    dataFrameFacturas.loc[index, 'Grupo Articulo'] = fila2['Grupo Artículo']
                    dataFrameFacturas.loc[index, 'Codigo de Grupo Articulo'] = fila2['Código de Grupo Artículo']
                    dataFrameFacturas.loc[index, 'Centro'] = fila2['Centro']
                    dataFrameFacturas.loc[index, 'Codigo de centro'] = fila2['Código de Centro']
                    dataFrameFacturas.loc[index, 'Servicio'] = fila2['Servicio']
                    dataFrameFacturas.loc[index, 'CECO'] = fila2['CECO']

    print(dataFrameFacturas)
    dataFrameFacturas.to_excel(rutaRepositorioOS + '11 Generado\\Facturas.xlsx', index=False)



def navegarSAP():
    print('\nComenzar a navegar por SAP para crear OS')
    dataFrame = pd.read_excel(rutaRepositorioOS + '11 Generado\\Facturas.xlsx')
    fechaFutura= datetime.now() + timedelta(5)
    fechaFutura = fechaFutura.strftime('%d.%m.%Y')
    print(fechaFutura)
    print('Archivo excel leído, se poseen: ', len(dataFrame), ' facturas')
    for index, fila in dataFrame.iterrows():
        dataFrame.loc[index, 'Número documento'] = ''
        if pd.isna(fila['Grupo Articulo']):
            print(fila['Detalle'], fila['Grupo Articulo'])

    # time.sleep(80)

    try:
        # colocamos la ruta de SAP
        Path ="C:\Program Files (x86)\SAP\FrontEnd\SAPgui\saplogon.exe"
        subprocess.Popen(Path)
        time.sleep(3)

        # Inicilizamos las instancias de SAP
        SapGuiAuto = win32com.client.GetObject('SAPGUI')
        if not type(SapGuiAuto)==win32com.client.CDispatch:
            return
        application = SapGuiAuto.GetScriptingEngine
        if not type(application)==win32com.client.CDispatch:
            SapGuiAuto=None
            return
        connection=application.Openconnection("3.02 - SAP QA2 - S/4",True)
        if not type(connection)==win32com.client.CDispatch:
            application=None
            SapGuiAuto=None
            return
        session=connection.children(0)
        if not type(session)==win32com.client.CDispatch:
            connection=None
            application=None
            SapGuiAuto=None
            return
        # Acá ingresamos nuestras credenciales, mandante y lenguaje.
        session.findById("wnd[0]/usr/txtRSYST-MANDT").Text = "100"
        session.findById("wnd[0]/usr/txtRSYST-BNAME").Text = user
        session.findById("wnd[0]/usr/pwdRSYST-BCODE").Text = password
        session.findById("wnd[0]/usr/txtRSYST-LANGU").Text = "ES"
        session.findById("wnd[0]").sendVKey(0) 
        session.findById("wnd[0]").sendVKey(0)
        session.findById("wnd[0]").maximize()
        session.findById("wnd[0]/tbar[0]/okcd").text = "ME21N"
        session.findById("wnd[0]").sendVKey(0)

        # Loopear por cada factura pra crear OC
        # contador = 13
        contador = 13
        for index, fila in dataFrame.iterrows():
            if pd.isna(fila['Grupo Articulo']) == False:#:s
                print('Índice: ', index)
                # time.sleep(80)
                print("wnd[0]/usr/subSUB0:SAPLMEGUI:00" + str(contador) + "/subSUB0:SAPLMEGUI:0030/subSUB1:SAPLMEGUI:1105/ctxtMEPO_TOPLINE-SUPERFIELD")
                session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:00" + str(contador) + "/subSUB0:SAPLMEGUI:0030/subSUB1:SAPLMEGUI:1105/ctxtMEPO_TOPLINE-SUPERFIELD").setFocus()
                session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:00" + str(contador) + "/subSUB0:SAPLMEGUI:0030/subSUB1:SAPLMEGUI:1105/ctxtMEPO_TOPLINE-SUPERFIELD").text = fila['Codigo Proveedor']
                # session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0013/subSUB0:SAPLMEGUI:0030/subSUB1:SAPLMEGUI:1105/ctxtMEPO_TOPLINE-SUPERFIELD").text
                session.findById("wnd[0]").sendVKey(0)
                print('--> Proveedor colocado')
                session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0013/subSUB1:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1102/tabsHEADER_DETAIL/tabpTABHDT8/ssubTABSTRIPCONTROL2SUB:SAPLMEGUI:1221/ctxtMEPO1222-EKORG").text = "PE11"
                session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0013/subSUB1:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1102/tabsHEADER_DETAIL/tabpTABHDT8/ssubTABSTRIPCONTROL2SUB:SAPLMEGUI:1221/ctxtMEPO1222-EKGRP").text = "340"
                session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0013/subSUB1:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1102/tabsHEADER_DETAIL/tabpTABHDT8/ssubTABSTRIPCONTROL2SUB:SAPLMEGUI:1221/ctxtMEPO1222-BUKRS").text = "PE11"
                session.findById("wnd[0]").sendVKey(27)
                session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0013/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1211/tblSAPLMEGUITC_1211/txtMEPO1211-EBELP[1,0]").text = "10" # Posicion
                session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0013/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1211/tblSAPLMEGUITC_1211/ctxtMEPO1211-KNTTP[2,0]").text = fila['Imputacion']
                session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0013/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1211/tblSAPLMEGUITC_1211/ctxtMEPO1211-EPSTP[3,0]").text = "F"
                print('--> a')
                session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0013/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1211/tblSAPLMEGUITC_1211/ctxtMEPO1211-EMATN[4,0]").text = ""
                print('--> Material')
                # print('-->', fila['Detalle'], ' ..')
                detalle = fila['Detalle'][:40]
                session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0013/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1211/tblSAPLMEGUITC_1211/txtMEPO1211-TXZ01[5,0]").text = detalle
                print('--> Detalle')
                session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0013/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1211/tblSAPLMEGUITC_1211/txtMEPO1211-MENGE[6,0]").text = "1"
                session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0013/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1211/tblSAPLMEGUITC_1211/ctxtMEPO1211-MEINS[7,0]").text = "SRV"
                session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0013/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1211/tblSAPLMEGUITC_1211/ctxtMEPO1211-ELPEI[8,0]").text = "D"
                session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0013/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1211/tblSAPLMEGUITC_1211/ctxtMEPO1211-EEIND[9,0]").text = fechaFutura
                print('-->', fechaFutura)
                monto = 0
                if fila['Incluye'] == 'Prima sin IGV':
                    monto = fila['Prima']
                else:
                    monto = fila['Prima total']

                try:
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0013/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1211/tblSAPLMEGUITC_1211/txtMEPO1211-NETPR[10,0]").text = monto
                except:
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:1303/tabsITEM_DETAIL/tabpTABIDT1/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1328/subSUB0:SAPLMLSP:0400/tblSAPLMLSPTC_VIEW/txtESLL-TBTWR[6,0]").text = monto
                # elif fila['Incluye'] == 'Prima más IGV':
                #     session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0013/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1211/tblSAPLMEGUITC_1211/txtMEPO1211-NETPR[10,0]").text = 
                # elif fila['Incluye'] == 'Total recargas':
                    # session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0013/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1211/tblSAPLMEGUITC_1211/txtMEPO1211-NETPR[10,0]").text = fila['Prima total']
                print('--> Prima')
                session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0013/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1211/tblSAPLMEGUITC_1211/txtMEPO1211-PEINH[12,0]").text = 1
                session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0013/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1211/tblSAPLMEGUITC_1211/ctxtMEPO1211-BPRME[13,0]").text = "SRV"
                session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0013/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1211/tblSAPLMEGUITC_1211/ctxtMEPO1211-WGBEZ[14,0]").text = fila['Grupo Articulo']
                session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0013/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1211/tblSAPLMEGUITC_1211/ctxtMEPO1211-NAME1[15,0]").text = fila['Codigo de centro']
                session.findById("wnd[0]").sendVKey(0)
                print('--> Por colocado')
                session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:1303/tabsITEM_DETAIL/tabpTABIDT1/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1328/subSUB0:SAPLMLSP:0400/tblSAPLMLSPTC_VIEW/ctxtESLL-SRVPOS[2,0]").text = fila['Servicio']
                session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:1303/tabsITEM_DETAIL/tabpTABIDT1/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1328/subSUB0:SAPLMLSP:0400/tblSAPLMLSPTC_VIEW/txtESLL-MENGE[4,0]").text = "1"
                session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:1303/tabsITEM_DETAIL/tabpTABIDT1/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1328/subSUB0:SAPLMLSP:0400/tblSAPLMLSPTC_VIEW/ctxtESLL-MEINS[5,0]").text = "SRV"
                monto = 0
                if fila['Incluye'] == 'Prima sin IGV':
                    monto = fila['Prima']
                else:
                    monto = fila['Prima total']
                try:
                # if fila['Incluye'] == 'Prima sin IGV':
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:1303/tabsITEM_DETAIL/tabpTABIDT1/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1328/subSUB0:SAPLMLSP:0400/tblSAPLMLSPTC_VIEW/txtESLL-TBTWR[6,0]").text = monto
                except:
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:1303/tabsITEM_DETAIL/tabpTABIDT1/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1328/subSUB0:SAPLMLSP:0400/tblSAPLMLSPTC_VIEW/txtESLL-TBTWR[6,0]").text = monto
                # elif fila['Incluye'] == 'Prima más IGV':
                    # session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:1303/tabsITEM_DETAIL/tabpTABIDT1/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1328/subSUB0:SAPLMLSP:0400/tblSAPLMLSPTC_VIEW/txtESLL-TBTWR[6,0]").text = fila['Prima total']
                # elif fila['Incluye'] == 'Total recargas':
                    # session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:1303/tabsITEM_DETAIL/tabpTABIDT1/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1328/subSUB0:SAPLMLSP:0400/tblSAPLMLSPTC_VIEW/txtESLL-TBTWR[6,0]").text = fila['Prima total']
                session.findById("wnd[0]").sendVKey(0)
                print('--> Enter')

                try: 
                    session.findById("wnd[1]/usr/subKONTBLOCK:SAPLKACB:1101/ctxtCOBL-PRCTR").text = fila['CECO']
                except:
                    session.findById("wnd[1]/usr/subKONTBLOCK:SAPLKACB:2100/ctxtCOBL-KOSTL").text = fila['CECO']
                print('--> CECO')
                session.findById("wnd[0]").sendVKey(0) # Enter
                session.findById("wnd[0]/tbar[0]/btn[11]").press() # Grabar
                session.findById("wnd[1]/usr/btnSPOP-VAROPTION1").press() # Confirmar paso de grabar
                numeroDocumento = str(session.findById("wnd[0]/sbar").text).split(' ')[6]
                print('Número de documentoc creado: ', numeroDocumento)
                dataFrame.loc[index, 'Número documento'] = numeroDocumento
                # print(dataFrame)
                dataFrame.to_excel(rutaRepositorioOS + '11 Generado\\Facturas.xlsx', index=False)
                # time.sleep(5)
            if index < 1:
                contador = contador + 3

    except Exception as e:
        mensaje = str(e)
        print('Error encontrado: ', mensaje)
        time.sleep(10)

        
print('Nombre de usuario: ', user)
print('Contraseña de usuario: ', password)
print('Ruta de repositorio: ', rutaRepositorio)
print('Ruta repositorio de MAPFRE: ', rutaRepositorioMapfre)
print('Ruta repositorio OS: ', rutaRepositorioOS)


# os.system("TASKKILL /F /IM saplogon.exe")




def crearHES():
    dataFrameFacturas = pd.read_excel(rutaRepositorioOS + '11 Generado\\Facturas.xlsx')
    print(dataFrameFacturas)
    # time.sleep(80)
    try:
        # colocamos la ruta de SAP
        Path ="C:\Program Files (x86)\SAP\FrontEnd\SAPgui\saplogon.exe"
        subprocess.Popen(Path)
        time.sleep(3)

        # Inicilizamos las instancias de SAP
        SapGuiAuto = win32com.client.GetObject('SAPGUI')
        if not type(SapGuiAuto)==win32com.client.CDispatch:
            return
        application = SapGuiAuto.GetScriptingEngine
        if not type(application)==win32com.client.CDispatch:
            SapGuiAuto=None
            return
        connection=application.Openconnection("3.02 - SAP QA2 - S/4",True)
        if not type(connection)==win32com.client.CDispatch:
            application=None
            SapGuiAuto=None
            return
        session=connection.children(0)
        if not type(session)==win32com.client.CDispatch:
            connection=None
            application=None
            SapGuiAuto=None
            return
        # Acá ingresamos nuestras credenciales, mandante y lenguaje.
        session.findById("wnd[0]/usr/txtRSYST-MANDT").Text = "100"
        session.findById("wnd[0]/usr/txtRSYST-BNAME").Text = user
        session.findById("wnd[0]/usr/pwdRSYST-BCODE").Text = password
        session.findById("wnd[0]/usr/txtRSYST-LANGU").Text = "ES"
        session.findById("wnd[0]").sendVKey(0) 
        session.findById("wnd[0]").sendVKey(0)
        session.findById("wnd[0]").maximize()
        session.findById("wnd[0]/tbar[0]/okcd").text = "ME21N"
        session.findById("wnd[0]").sendVKey(0)

        # Loopear por cada factura pra crear OC
        for index, fila in dataFrameFacturas.iterrows():
            if not pd.isna(fila['Número documento']):
                print('Número de documento: ', fila['Número documento'], index)
                # time.sleep(80)
                # time.sleep(800)
                session.findById("wnd[0]").sendVKey(17)
                session.findById("wnd[0]").maximize
                session.findById("wnd[0]/tbar[1]/btn[17]").press()
                session.findById("wnd[1]/usr/subSUB0:SAPLMEGUI:0003/ctxtMEPO_SELECT-EBELN").text = fila['Número documento']
                session.findById("wnd[1]/usr/subSUB0:SAPLMEGUI:0003/radMEPO_SELECT-BSTYP_F").setFocus()
                session.findById("wnd[1]/usr/subSUB0:SAPLMEGUI:0003/radMEPO_SELECT-BSTYP_F").select()
                session.findById("wnd[1]").sendVKey(0)
                session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB1:SAPLMEVIEWS:1100/subSUB1:SAPLMEVIEWS:4000/btnDYN_4000-BUTTON").press()
                session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0013/subSUB1:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1102/tabsHEADER_DETAIL/tabpTABHDT10").select()
                session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0013/subSUB1:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1102/tabsHEADER_DETAIL/tabpTABHDT10/ssubTABSTRIPCONTROL2SUB:SAPLMERELVI:1100/txtMEPO_REL_GENERAL-FRGET").setFocus
                session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0013/subSUB1:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1102/tabsHEADER_DETAIL/tabpTABHDT10/ssubTABSTRIPCONTROL2SUB:SAPLMERELVI:1100/txtMEPO_REL_GENERAL-FRGET").caretPosition = 3
                estado = session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0013/subSUB1:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1102/tabsHEADER_DETAIL/tabpTABHDT10/ssubTABSTRIPCONTROL2SUB:SAPLMERELVI:1100/txtMEPO_REL_GENERAL-FRGET").text
                if estado == 'Liberado':
                    print('Estado: ', estado)
                    session.findById("wnd[0]").sendVKey(28)
                    # session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:1303/tabsITEM_DETAIL/tabpTABIDT1/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1328/subSUB0:SAPLMLSP:0400/tblSAPLMLSPTC_VIEW").verticalScrollbar.position = 3
                    # session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:1303/tabsITEM_DETAIL/tabpTABIDT1/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1328/subSUB0:SAPLMLSP:0400/tblSAPLMLSPTC_VIEW/ctxtESLL-SRVPOS[2,0]").caretPosition = 7
                    # time.sleep(8)
                    # print('Hola ')
                    # session.findById("wnd[0]").sendVKey(28)
                    session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:1303/tabsITEM_DETAIL/tabpTABIDT1").select()
                    numeroServicio = session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:1303/tabsITEM_DETAIL/tabpTABIDT1/ssubTABSTRIPCONTROL1SUB:SAPLMEGUI:1328/subSUB0:SAPLMLSP:0400/tblSAPLMLSPTC_VIEW/ctxtESLL-SRVPOS[2,0]").text
                    # numeroServicio = session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0013/subSUB1:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1102/tabsHEADER_DETAIL/tabpTABHDT10/ssubTABSTRIPCONTROL2SUB:SAPLMERELVI:1100/cntlRELEASE_INFO/shellcont/shell").text
                    print(numeroServicio)
                    # time.sleep(80)
                    print('Número de servicio: ', numeroServicio)
                    # time.sleep(80000)
                    session.findById("wnd[0]").sendVKey(3)

                    # Entrada de órdenes
                    session.findById("wnd[0]/usr/cntlIMAGE_CONTAINER/shellcont/shell/shellcont[0]/shell").topNode = ("0000000026")
                    session.findById("wnd[0]/usr/cntlIMAGE_CONTAINER/shellcont/shell/shellcont[0]/shell").doubleClickNode("0000000238")
                    session.findById("wnd[0]").sendVKey(17)
                    session.findById("wnd[1]/usr/ctxtRM11R-EBELN").text = fila['Número documento']
                    session.findById("wnd[1]/tbar[0]/btn[0]").press()
                    session.findById("wnd[0]/tbar[1]/btn[13]").press()
                    session.findById("wnd[0]/usr/txtESSR-TXZ01").text = fila['Detalle']
                    session.findById("wnd[0]/usr/subSERVICE:SAPLMLSP:0400/tblSAPLMLSPTC_VIEW/txtESLL-MENGE[4,0]").text = "1"
                    session.findById("wnd[0]/usr/subSERVICE:SAPLMLSP:0400/tblSAPLMLSPTC_VIEW/ctxtESLL-MEINS[5,0]").text = "SRV"
                    if fila['Incluye'] == 'Prima sin IGV':
                        session.findById("wnd[0]/usr/subSERVICE:SAPLMLSP:0400/tblSAPLMLSPTC_VIEW/txtESLL-TBTWR[6,0]").text = fila['Prima']
                    elif fila['Incluye'] == 'Prima más IGV':
                        session.findById("wnd[0]/usr/subSERVICE:SAPLMLSP:0400/tblSAPLMLSPTC_VIEW/txtESLL-TBTWR[6,0]").text = fila['Prima total']
                    elif fila['Incluye'] == 'Total recargas':
                        session.findById("wnd[0]/usr/subSERVICE:SAPLMLSP:0400/tblSAPLMLSPTC_VIEW/txtESLL-TBTWR[6,0]").text = fila['Prima total']

                    session.findById("wnd[0]/usr/subSERVICE:SAPLMLSP:0400/tblSAPLMLSPTC_VIEW/ctxtESLL-SRVPOS[2,0]").text = numeroServicio
                    session.findById("wnd[0]/usr/subSERVICE:SAPLMLSP:0400/tblSAPLMLSPTC_VIEW/txtESLL-TBTWR[6,0]").setFocus
                    session.findById("wnd[0]/usr/subSERVICE:SAPLMLSP:0400/tblSAPLMLSPTC_VIEW/txtESLL-TBTWR[6,0]").caretPosition = 8

                    # time.sleep(80)
                    session.findById("wnd[0]").sendVKey(0)
                    session.findById("wnd[0]/tbar[0]/btn[11]").press()

                    session.findById("wnd[1]/usr/btnSPOP-OPTION1").press()
                    numeroHES = session.findById("wnd[0]/usr/ctxtESSR-LBLNI").text
                    print('Número HES: ', numeroHES, '\n')
                    dataFrameFacturas.loc[index, 'HES'] = numeroHES
                    session.findById("wnd[0]").sendVKey(3)
                    session.findById("wnd[0]/tbar[0]/okcd").text = "ME21N"
                    session.findById("wnd[0]").sendVKey(0)
                else:
                    print('No liberado')
                # time.sleep(1)
            else:
                print(fila['Detalle'], ' no tiene número de documento')
        dataFrameFacturas.to_csv(rutaRepositorio + 'Facturas - HES.csv', index = False)
        print('Se han creado las HES')

    except Exception as e:
        mensaje = str(e)
        print('Error encontrado: ', mensaje)
        time.sleep(10)

def crearTercerCodigo():
    print('Comenzar a capturar 3er código y crear anticipo de ser necesario')
    rutaRepositorio = 'D:/Repository-Emision de OS/Exportado/'
    dataFrame = pd.read_csv(rutaRepositorioOS + '11 Generado\\Facturas - HES.csv')
    try:
        # colocamos la ruta de SAP
        Path ="C:\Program Files (x86)\SAP\FrontEnd\SAPgui\saplogon.exe"
        subprocess.Popen(Path)
        time.sleep(3)

        # Inicilizamos las instancias de SAP
        SapGuiAuto = win32com.client.GetObject('SAPGUI')
        if not type(SapGuiAuto)==win32com.client.CDispatch:
            return
        application = SapGuiAuto.GetScriptingEngine
        if not type(application)==win32com.client.CDispatch:
            SapGuiAuto=None
            return
        connection=application.Openconnection("3.02 - SAP QA2 - S/4",True)
        if not type(connection)==win32com.client.CDispatch:
            application=None
            SapGuiAuto=None
            return
        session=connection.children(0)
        if not type(session)==win32com.client.CDispatch:
            connection=None
            application=None
            SapGuiAuto=None
            return
        # Acá ingresamos nuestras credenciales, mandante y lenguaje.
        session.findById("wnd[0]/usr/txtRSYST-MANDT").Text = "100"
        session.findById("wnd[0]/usr/txtRSYST-BNAME").Text = user
        session.findById("wnd[0]/usr/pwdRSYST-BCODE").Text = password
        session.findById("wnd[0]/usr/txtRSYST-LANGU").Text = "ES"
        session.findById("wnd[0]").sendVKey(0) 
        session.findById("wnd[0]").sendVKey(0)

        for index, fila in dataFrame.iterrows():
            if not pd.isna(fila['HES']):
                print('Número de documento ',fila['Número documento'], ' con HES: ', fila['HES'])
                session.findById("wnd[0]/usr/cntlIMAGE_CONTAINER/shellcont/shell/shellcont[0]/shell").expandNode("0000000034")
                session.findById("wnd[0]/usr/cntlIMAGE_CONTAINER/shellcont/shell/shellcont[0]/shell").selectedNode = "0000000246"
                session.findById("wnd[0]/usr/cntlIMAGE_CONTAINER/shellcont/shell/shellcont[0]/shell").topNode = "0000000026"
                session.findById("wnd[0]/usr/cntlIMAGE_CONTAINER/shellcont/shell/shellcont[0]/shell").doubleClickNode("0000000246")
                session.findById("wnd[0]/tbar[1]/btn[17]").press()
                session.findById("wnd[1]/usr/subSUB0:SAPLMEGUI:0003/ctxtMEPO_SELECT-EBELN").text = fila['Número documento']
                session.findById("wnd[1]").sendVKey(0)

                os.system("TASKKILL /F /IM excel.exe")
                session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:1303/tabsITEM_DETAIL/tabpTABIDT3").select()
                session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:1303/tabsITEM_DETAIL/tabpTABIDT4").select()
                session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0015/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:1303/tabsITEM_DETAIL/tabpTABIDT16").select()
                session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:1303/tabsITEM_DETAIL/tabpTABIDT16/ssubTABSTRIPCONTROL1SUB:SAPLMEVIEWS:1106/subSUB2:SAPLMEGUI:1316/ssubPO_HISTORY:SAPLMMHIPO:0100/cntlMEALV_GRID_CONTROL_MMHIPO/shellcont/shell").pressToolbarContextButton("&MB_EXPORT")
                session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:1303/tabsITEM_DETAIL/tabpTABIDT16/ssubTABSTRIPCONTROL1SUB:SAPLMEVIEWS:1106/subSUB2:SAPLMEGUI:1316/ssubPO_HISTORY:SAPLMMHIPO:0100/cntlMEALV_GRID_CONTROL_MMHIPO/shellcont/shell").selectContextMenuItem("&XXL")
                session.findById("wnd[1]/usr/subSUB_CONFIGURATION:SAPLSALV_GUI_CUL_EXPORT_AS:0512/txtGS_EXPORT-FILE_NAME").text = "archivo"
                session.findById("wnd[1]/usr/subSUB_CONFIGURATION:SAPLSALV_GUI_CUL_EXPORT_AS:0512/cmbGS_EXPORT-FORMAT").setFocus
                session.findById("wnd[1]/tbar[0]/btn[20]").press()
                session.findById("wnd[1]/usr/ctxtDY_PATH").text = rutaRepositorio
                session.findById("wnd[1]/usr/ctxtDY_FILENAME").text = "archivo.xlsx"
                session.findById("wnd[1]/tbar[0]/btn[11]").press()
                time.sleep(2)
                os.system("TASKKILL /F /IM excel.exe")
                dataFrameAuxiliar = pd.read_excel(rutaRepositorio + 'archivo.xlsx')
                # print('Checkpoint')
                
                # time.sleep(80)


                # print('eje')
                if dataFrameAuxiliar.loc[0, 'Txt.brv.'] == 'WE':
                    tercerCodigo = str(dataFrameAuxiliar.loc[0, 'Documento material']).replace('.0', '')
                    print('Tercer código: ', tercerCodigo)
                    dataFrame.loc[index, 'Aprobacion HES'] = tercerCodigo
                    print('¿Tiene el número de documento anticipo asociado? ', fila['Anticipo'])
                    if fila['Anticipo'] == 'SI':
                        print('--Añadir anticipo')
                        session.findById("wnd[0]/tbar[0]/btn[3]").press()
                        session.findById("wnd[0]/tbar[0]/okcd").text = 'F-47'
                        session.findById("wnd[0]").sendVKey(0) 
                        fechaActual = datetime.now().strftime('%d.%m.%Y')
                        session.findById("wnd[0]/usr/ctxtBKPF-BLDAT").text = fechaActual 
                        session.findById("wnd[0]/usr/ctxtBKPF-BUKRS").text = fila['Sociedad']
                        session.findById("wnd[0]/usr/txtBKPF-MONAT").text = fechaActual.split('.')[1]
                        session.findById("wnd[0]/usr/ctxtBKPF-WAERS").text = "PEN"
                        detalle = str(fila['Detalle']).lstrip()
                        # print(detalle[:16])
                        session.findById("wnd[0]/usr/txtBKPF-BKTXT").text = detalle
                        session.findById("wnd[0]/usr/ctxtRF05A-ZUMSK").text = "A"
                        session.findById("wnd[0]/usr/ctxtRF05A-NEWKO").text = fila['Codigo Proveedor']
                        session.findById("wnd[0]/usr/txtBKPF-XBLNR").text = detalle[:16]
                        # print('uwu')
                        session.findById("wnd[0]").sendVKey(0)
                        session.findById("wnd[0]/usr/txtBSEG-WRBTR").text = str(fila['Prima total']).replace(',', '')
                        session.findById("wnd[0]/usr/ctxtBSEG-ZFBDT").text = fechaActual
                        session.findById("wnd[0]/usr/ctxtBSEG-ZLSCH").text = "T"
                        session.findById("wnd[0]/usr/ctxtBSEG-EBELN").text = fila['Número documento']
                        session.findById("wnd[0]/usr/ctxtBSEG-EBELP").text = "10"
                        session.findById("wnd[0]/usr/ctxtBSEG-SGTXT").text = fila['Detalle']
                        session.findById("wnd[0]/tbar[0]/btn[11]").press()
                        session.findById("wnd[0]").sendVKey(0)
                        session.findById("wnd[0]").sendVKey(0)
                        textoCompleto = session.findById("wnd[0]/sbar").text
                        numeroAnticipo = textoCompleto.split(' ')[0] 
                        numeroAnticipo = numeroAnticipo.split('.')[1]
                        print('Número de anticipo: ', numeroAnticipo)
                        dataFrame.loc[index, 'Numero de anticipo'] = numeroAnticipo
                        # session.findById("wnd[0]/sbar").doubleClick()
                        time.sleep(1)
                        session.findById("wnd[0]/tbar[0]/btn[3]").press()
                        session.findById("wnd[1]/usr/btnSPOP-OPTION1").press()

                    else:
                        session.findById("wnd[0]/tbar[0]/btn[3]").press()
                    dataFrame.to_csv(rutaRepositorio + 'Facturas - Anticipos.csv', index = False)
                    
            else:
                print('Número de documento ',fila['Número documento'], ' no tiene HES')

    except Exception as e:
        print('Error encontrado: ', e)
    










# dataFrameNuevosNombres = leerNuevosNombres()
dataFrameFacturas = leerCarpetas()
# dataFrameFacturasMAPFRE = leerCarpetaMAPFRE()
# dataFrameOncocenter = leerONCOCENTER()
# dataFrameTEBCA = leerTEBCA() 
# adicionarInformacion(dataFrameFacturas, dataFrameFacturasMAPFRE, dataFrameOncocenter, dataFrameTEBCA)
# leerGrupoArticulo()
os.system("TASKKILL /F /IM saplogon.exe")
# navegarSAP()


# os.system("TASKKILL /F /IM excel.exe")
# crearHES()

# crearTercerCodigo()
# os.system("TASKKILL /F /IM saplogon.exe")


