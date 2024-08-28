import os
from dataclasses import dataclass
from datetime import datetime, timedelta
import PyPDF2
import pandas as pd
import time
from pdfminer.high_level import extract_text
import re
import openpyxl 
import win32com.client
import subprocess
import sys
# import pdb; pdb.set_trace()

original_sys = sys.stdout

@dataclass
class claseFactura:
    sociedad : str
    asegurado : str
    grupoPersonal : str
    detalle : str
    prima : str
    igv : str
    primaTotal : str

# Definir la función limpiar y convertir 
def limpiar_y_convertir(cadena):
    # Remover cualquier carácter no numérico excepto el punto decimal
    cadena_limpia = re.sub(r'[^\d.]', '', cadena)
    try:
        return float(cadena_limpia)
    except ValueError:
        return None  # O algún valor por defecto o lanzar una excepción
    

hojaDatos = openpyxl.load_workbook('90 Excel Config.xlsx')
workSheet = hojaDatos.active

listaFacturas = []
fechaActual = datetime.now()
fechaActualTexto = fechaActual.strftime('%d.%m.%Y')
fechaSiguiente = fechaActual + timedelta(7)
fechaSiguienteTexto = fechaSiguiente.strftime('%d.%m.%Y')

user = workSheet.cell(row = 6, column = 6).value
password = workSheet.cell(row = 7, column = 6).value
rutaRepositorio = workSheet.cell(row = 8, column = 6).value
rutaRepositorioMapfre = workSheet.cell(row = 9, column = 6).value
rutaRepositorioOS = workSheet.cell(row = 10, column = 6).value

hojaDatos.close()

dataFrameNuevosNombres = None
dataFrameSociedades = pd.read_excel(rutaRepositorioOS + 'Sociedades.xlsx')
print(dataFrameSociedades)

def leerNuevosNombres():
    dataFrameAuxiliar = pd.read_excel(rutaRepositorioOS + 'Nuevos nombres.xlsx')
    for index, fila in dataFrameAuxiliar.iterrows():
        nuevoNombre = fila['DETALLE'].split(' ')
        nuevoNombre[len(nuevoNombre) - 1] = datetime.now().strftime('%m.%y')
        nuevoDetalle = ''
        for elemento in nuevoNombre:
            nuevoDetalle = nuevoDetalle + ' ' + elemento + ' '
        nuevoDetalle = nuevoDetalle.replace('  ', ' ')
        dataFrameAuxiliar.loc[index, 'DETALLE'] = nuevoDetalle
    return dataFrameAuxiliar

def leerCarpetas():
    print(rutaRepositorio)
    listaCarpetas = os.listdir(rutaRepositorio)
    print('\nComenzar a revisar información de facturas PDF en carpeta')
    for carpeta in listaCarpetas:
        if carpeta.endswith('.xlsx') == False and carpeta.endswith('.csv') == False:
            listaSubCarpetas = os.listdir(rutaRepositorio + '\\'+ carpeta)
            for archivo in listaSubCarpetas:
                if archivo.endswith('.pdf'):
                    nombreSociedad = archivo.split('-')[0]
                    archivoPDF = open(rutaRepositorio + '\\'+ carpeta + '\\' + archivo, 'rb')
                    lectorPDF = PyPDF2.PdfReader(archivoPDF)
                    cantidadPaginas = len(lectorPDF.pages)
                    print('Cantidad: ', cantidadPaginas, archivo, '===================================')
                    try:
                        for paginaNumero in range(cantidadPaginas):
                            pagina = lectorPDF.pages[paginaNumero]
                            texto = pagina.extract_text()
                            texto = texto.split('\n')
                            # for indice in range(len(texto)):
                            #     print(indice, texto[indice])
                            if texto[14].startswith('Asesor: ') == False:
                                prima = limpiar_y_convertir(texto[33])
                                igv = limpiar_y_convertir(texto[36])
                                primaTotal = limpiar_y_convertir(texto[29])
                                print('primar: ',prima,' igv: ',igv,' prima total: ',primaTotal)
                            else:
                                prima = limpiar_y_convertir(texto[34])
                                igv = limpiar_y_convertir(texto[37])
                                primaTotal = limpiar_y_convertir(texto[30])
                                print('primar: ',prima,' igv: ',igv,' prima total: ',primaTotal)
                            if prima is None or igv is None or primaTotal is None:
                                print(f"Error encontrado en archivo {archivo}: No se pudo convertir alguno de los valores.")
                                continue  # Salta al siguiente archivo si hay un error

                            nombreArchivo = archivo.replace('.pdf', '')
                            listaFacturas.append(claseFactura(nombreSociedad, '', '', nombreArchivo, prima, igv, primaTotal))
                    except Exception as e:
                        print('Error encontrado: ', e)
                        time.sleep(8000)
    
    dataFrameAuxiliar = pd.DataFrame(
        [(factura.sociedad, factura.detalle, factura.prima, factura.igv, factura.primaTotal) for factura in listaFacturas],
        columns=['Sociedad', 'nombreArchivo', 'prima', 'igv', 'primaTotal']
    )
    print(dataFrameAuxiliar)
    return dataFrameAuxiliar

def leerCarpetaMAPFRE():
    print('Comenzar a leer información de facturas de MAPFRE:\n')
    listaFacturas = os.listdir(rutaRepositorioMapfre)
    listaFacturasLeidas = []
    for archivo in listaFacturas:
        if archivo.endswith('.PDF') or archivo.endswith('.pdf'):
            archivoPDF = open(rutaRepositorioMapfre + archivo, 'rb')
            lectorPDF = PyPDF2.PdfReader(archivoPDF)
            cantidadPaginas = len(lectorPDF.pages)
            try:
                texto = extract_text(rutaRepositorioMapfre + archivo)
                texto = texto.split('\n')
                for paginaNumero in range(cantidadPaginas):
                    pagina = lectorPDF.pages[paginaNumero]
                    texto = pagina.extract_text()
                    texto = texto.split('\n')
                    igv = 0
                    prima = 0
                    primaTotal = 0
                    for numeroElemento in range(len(texto)):
                        if texto[numeroElemento].startswith('Prima Comercial + IGV :'):
                            linea = texto[numeroElemento].split(' ')
                            primaTotal = float(linea[6].replace(',', ''))
                            # print('Prima total: ', primaTotal, ' archivo: ', archivo)
                            prima = primaTotal / 1.18
                            igv = prima * 0.18
                    if texto[0].startswith('VIDA'):
                        if texto[38].startswith('EMPLEADOS'):
                            detalle = 'VIDA LEY EMPLEADOS'
                            if texto[28].startswith('ALICORP') == False:
                                grupoPersonal = 'CESADO'
                            else:
                                grupoPersonal = 'EMPLEADO'
                        else:
                            detalle = 'VIDA LEY OBREROS'
                        contratante = texto[28]
                    else:
                        for numeroElemento in range(len(texto)):
                            indiceAlterno = 0
                            if texto[numeroElemento].startswith('GRUPO MAPFRE PERUEl presente documento no'):
                                indiceAlterno = numeroElemento - 2
                                detalle = texto[indiceAlterno].replace('/', '')
                                detalle = re.sub(r"\d+", "", detalle)
                                detalle = detalle.strip()
                                detalle = detalle.replace('SEGURO COMPLEMENTARIO DE TRABAJO DE RIESGO', 'SCTR')
                                detalle = detalle.replace('PENSIONES', 'PENSIÓN')

                        if texto[27].startswith('ALICORP') == False:
                            grupoPersonal = 'CESADO'
                        else:
                            grupoPersonal = 'EMPLEADO'
                        contratante = texto[27]
                    detalle = detalle + ' ' + datetime.now().strftime('%m.%y')
                    # Revisar sociedad en dataframe de facturas
                    for index, fila in dataFrameSociedades.iterrows():
                        if contratante.upper().startswith(fila['Clave'].upper()):
                            sociedad = fila['CÓDIGO']
                        
                listaFacturasLeidas.append(claseFactura(sociedad, grupoPersonal, '', detalle, prima, igv, primaTotal))
            except Exception as e:
                print('Error encontrado: ', str(e))
    dataFrameAuxiliar = pd.DataFrame(
        [(factura.sociedad, factura.detalle, factura.prima, factura.igv, factura.primaTotal) for factura in listaFacturasLeidas],
        columns=['Sociedad', 'nombreArchivo', 'prima', 'igv', 'primaTotal']
    )
    print(dataFrameAuxiliar)
    return dataFrameAuxiliar       

def leerONCOCENTER():
    print('Comenzar a leer facturas de oncocenter')
    listaArchivos =  os.listdir(rutaRepositorioOS + 'ONCOCENTER\\')
    listaFacturas = []
    for archivo in listaArchivos:
        if archivo.upper().endswith('PDF'):
            texto = extract_text(rutaRepositorioOS + 'ONCOCENTER\\' + archivo, 'rb')
            texto = texto.split('\n')
            try:
                prima = float(texto[135].replace(',',''))
                igv = prima * 0.18
                primaTotal = prima * 1.18
                listaFacturas.append(claseFactura('PE11', '', 'EMPLEADO', 'Chequeo médico ' + datetime.now().strftime('%m.%y'), prima, igv, primaTotal))
            except Exception as e:
                print('Error encontrado: ', e)
    dataFrameAuxiliar = pd.DataFrame(
        [(factura.sociedad, factura.detalle, factura.prima, factura.igv, factura.primaTotal) for factura in listaFacturas],
        columns=['Sociedad', 'nombreArchivo', 'prima', 'igv', 'primaTotal']
    ) 
    return dataFrameAuxiliar

def leerTEBCA():
    print('Comenzar a leer facturas de TEBCA:')
    listaArchivos =  os.listdir(rutaRepositorioOS + 'TEBCA\\COMISION\\')
    listaFacturas = []
    for archivo in listaArchivos:
        if archivo.upper().endswith('PDF'):
            texto = extract_text(rutaRepositorioOS + 'TEBCA\\COMISION\\' + archivo, 'rb')
            texto = texto.split('\n')
            try:
                prima = float(texto[88].replace(',','').replace('S/', ''))
                igv = prima * 0.18
                primaTotal = prima * 1.18
                listaFacturas.append(claseFactura('PE14', '', 'EMPLEADO', 'RECARGA VALES COMISION ' + datetime.now().strftime('%m.%y'), prima, igv, primaTotal))
            except Exception as e:
                print('Error encontrado: ', e)

    listaArchivos =  os.listdir(rutaRepositorioOS + 'TEBCA\\RECARGA VALES VISA\\')
    for archivo in listaArchivos:
        if archivo.upper().endswith('PDF'):
            texto = extract_text(rutaRepositorioOS + 'TEBCA\\RECARGA VALES VISA\\' + archivo, 'rb')
            texto = texto.split('\n')
            try:
                prima = float(texto[97].replace(',','').replace('S/', ''))
                igv = prima * 0.18
                primaTotal = prima * 1.18
                listaFacturas.append(claseFactura('PE14', '', 'EMPLEADO', 'RECARGA VALES VISA ' + datetime.now().strftime('%m.%y'), prima, igv, primaTotal))
            except Exception as e:
                print('Error encontrado: ', e)
    dataFrameAuxiliar = pd.DataFrame(
        [(factura.sociedad, factura.detalle, factura.prima, factura.igv, factura.primaTotal) for factura in listaFacturas],
        columns=['Sociedad', 'nombreArchivo', 'prima', 'igv', 'primaTotal']
    )
    return dataFrameAuxiliar

def adicionarInformacion(dataFrameFacturas, dataFrameFacturasMAPFRE, dataFrameOncocenter, dataFrameTEBCA):
    print('\nComenzar a modificar información de dataFrame')
    
    def normalizar_texto(texto):
        # Convertir a mayúsculas, eliminar espacios en blanco y normalizar la cadena
        return re.sub(r'\s+', '', str(texto).strip().upper())
    
    # MAPFRE
    print('MAPFRE: ')
    for index, fila in dataFrameFacturasMAPFRE.iterrows():
        nombreArchivoAuxiliar = normalizar_texto(fila['nombreArchivo'])
        for index2, fila2 in dataFrameNuevosNombres.iterrows():
            detalleAuxiliar = normalizar_texto(fila2['DETALLE'])
            if nombreArchivoAuxiliar == detalleAuxiliar:
                dataFrameFacturasMAPFRE.loc[index, 'nombreArchivo'] = fila2['DETALLE']
                dataFrameFacturasMAPFRE.loc[index, 'Imputacion'] = fila2['I']
                dataFrameFacturasMAPFRE.loc[index, 'Grupo'] = fila2['Grupo de Personal']
                dataFrameFacturasMAPFRE.loc[index, 'Anticipo'] = fila2['ANTICIPO']
                dataFrameFacturasMAPFRE.loc[index, 'Incluye'] = fila2['Incluye']
                dataFrameFacturasMAPFRE.loc[index, 'Proveedor'] = fila2['Proveedor']
                dataFrameFacturasMAPFRE.loc[index, 'Codigo Proveedor'] = fila2['Cod. Proveedor']
                break  # Salir del bucle cuando se encuentre una coincidencia
    print('Data adicional de facturas de MAPFRE añadida')
    
    # PACIFICO
    print('PACIFICO: ')
    for index, fila in dataFrameFacturas.iterrows():
        nombreArchivoAuxiliar = normalizar_texto('-'.join(str(fila['nombreArchivo']).split('-')[1:]))
        for index2, fila2 in dataFrameNuevosNombres.iterrows():
            if fila2['Proveedor'] == 'PACIFICO':
                distintivoAuxiliar = normalizar_texto('-'.join(str(fila2['DISTINTIVO FACT']).split('-')[1:]))
                if nombreArchivoAuxiliar == distintivoAuxiliar:
                    dataFrameFacturas.loc[index, 'nombreArchivo'] = fila2['DETALLE']
                    dataFrameFacturas.loc[index, 'Imputacion'] = fila2['I']
                    dataFrameFacturas.loc[index, 'Grupo'] = fila2['Grupo de Personal']
                    dataFrameFacturas.loc[index, 'Anticipo'] = fila2['ANTICIPO']
                    dataFrameFacturas.loc[index, 'Incluye'] = fila2['Incluye']
                    dataFrameFacturas.loc[index, 'Proveedor'] = fila2['Proveedor']
                    dataFrameFacturas.loc[index, 'Codigo Proveedor'] = fila2['Cod. Proveedor']
                    break  # Salir del bucle cuando se encuentre una coincidencia
    print('Data adicional de facturas de PACIFICO añadida')

    # ONCOCENTER
    print('ONCOCENTER: ')
    for index, fila in dataFrameOncocenter.iterrows():
        dataFrameOncocenter.loc[index, 'nombreArchivo'] = 'Chequeo médico ' + datetime.now().strftime('%m.%y')
        dataFrameOncocenter.loc[index, 'Imputacion'] = 'K'
        dataFrameOncocenter.loc[index, 'Grupo'] = 'EMPLEADO'
        dataFrameOncocenter.loc[index, 'Anticipo'] = 'NO'
        dataFrameOncocenter.loc[index, 'Incluye'] = 'Prima sin IGV'   
        dataFrameOncocenter.loc[index, 'Proveedor'] = 'ONCOCENTER PERÚ'
        dataFrameOncocenter.loc[index, 'Codigo Proveedor'] = 1000002873
    print('Data adicional de facturas de ONCOCENTER añadida')

    # TEBCA
    print('TEBCA: ')
    for index, fila in dataFrameTEBCA.iterrows():
        if fila['nombreArchivo'] == 'RECARGA VALES VISA ' + datetime.now().strftime('%m.%y'):
            dataFrameTEBCA.loc[index, 'Imputacion'] = 'H'
            dataFrameTEBCA.loc[index, 'Anticipo'] = 'SI'
            dataFrameTEBCA.loc[index, 'Incluye'] = 'Total Recargas' 
        else:
            dataFrameTEBCA.loc[index, 'Imputacion'] = 'K'
            dataFrameTEBCA.loc[index, 'Anticipo'] = 'NO'
            dataFrameTEBCA.loc[index, 'Incluye'] = 'Prima sin IGV'   
        dataFrameTEBCA.loc[index, 'Grupo'] = 'EMPLEADO'
        dataFrameTEBCA.loc[index, 'Proveedor'] = 'TEBCA'
        dataFrameTEBCA.loc[index, 'Codigo Proveedor'] = 1000002873
    print('Data adicional de facturas de TEBCA añadida')

    # Cambiar nombre de columna y guardar
    dataFrameFacturas = dataFrameFacturas.rename(columns={'nombreArchivo': 'Detalle'})
    dataFrameFacturasMAPFRE = dataFrameFacturasMAPFRE.rename(columns={'nombreArchivo': 'Detalle'})
    dataFrameOncocenter = dataFrameOncocenter.rename(columns={'nombreArchivo': 'Detalle'})
    dataFrameTEBCA = dataFrameTEBCA.rename(columns={'nombreArchivo': 'Detalle'})

    dataFrameAuxiliar = pd.concat([dataFrameFacturasMAPFRE, dataFrameFacturas, dataFrameOncocenter, dataFrameTEBCA], ignore_index=True)
    dataFrameAuxiliar = dataFrameAuxiliar.rename(columns={'prima': 'Prima', 'primaTotal': 'Prima total', 'igv': 'IGV'})

    for index, fila in dataFrameAuxiliar.iterrows():
        dataFrameAuxiliar.loc[index, 'Prima total'] = "{:,.2f}".format(fila['Prima total'])
        dataFrameAuxiliar.loc[index, 'Prima'] = "{:,.2f}".format(fila['Prima'])
        dataFrameAuxiliar.loc[index, 'IGV'] = "{:,.2f}".format(fila['IGV'])

    print('\nDataFrame combinado: ')
    dataFrameAuxiliar.to_excel(rutaRepositorioOS + '11 Generado\\Facturas.xlsx', index=False)
    print('Se ha actualizado archivo de excel')


def leerGrupoArticulo():
    dataFrameGrupo = pd.read_excel(rutaRepositorioOS + 'Grupo articulo y varios.xlsx')
    # print(dataFrameGrupo)
    dataFrameFacturas = pd.read_excel(rutaRepositorioOS + '11 Generado\\Facturas.xlsx')

    for index, fila in dataFrameFacturas.iterrows():
        dataFrameFacturas.loc[index, 'Grupo Articulo'] = ''
        dataFrameFacturas.loc[index, 'Codigo de Grupo Articulo'] = ''
        dataFrameFacturas.loc[index, 'Centro'] = ''
        dataFrameFacturas.loc[index, 'Codigo de centro'] = ''
        dataFrameFacturas.loc[index, 'Servicio'] = ''
        dataFrameFacturas.loc[index, 'CECO'] = ''

    for index, fila in dataFrameFacturas.iterrows():
        for index2, fila2 in dataFrameGrupo.iterrows():
            # print(fila2['DETALLE'])
            if fila['Sociedad'] == fila2['Codigo de Sociedad']:
                detalleAuxiliar = str(fila['Detalle']).split()
                detalleAuxiliar = ''.join(detalleAuxiliar[:-1])

                detalleAuxuliarGrupo = str(fila2['DETALLE']).replace(' ', '')
                detalleAuxuliarGrupo = ''.join(detalleAuxuliarGrupo)

                print(detalleAuxiliar, detalleAuxuliarGrupo)
                # print(detalleAuxiliar, detalleAuxuliarGrupo, index2, fila['Sociedad'], fila2['Codigo de Sociedad'])
                if detalleAuxiliar == detalleAuxuliarGrupo:
                    fila['Detalle'].startswith(fila2['DETALLE'])
                    dataFrameFacturas.loc[index, 'Grupo Articulo'] = fila2['Grupo Artículo']
                    dataFrameFacturas.loc[index, 'Codigo de Grupo Articulo'] = fila2['Código de Grupo Artículo']
                    dataFrameFacturas.loc[index, 'Centro'] = fila2['Centro']
                    dataFrameFacturas.loc[index, 'Codigo de centro'] = fila2['Código de Centro']
                    dataFrameFacturas.loc[index, 'Servicio'] = fila2['Servicio']
                    dataFrameFacturas.loc[index, 'CECO'] = fila2['CECO']

    #####################################
    print(dataFrameFacturas)
    # dataFrameFacturas.to_excel(rutaRepositorioOS + '11 Generado\\Facturas.xlsx', index=False)

# dataFrameNuevosNombres = leerNuevosNombres()
# dataFrameFacturas = leerCarpetas()
# dataFrameFacturasMAPFRE = leerCarpetaMAPFRE()
# dataFrameOncocenter = leerONCOCENTER()
# dataFrameTEBCA = leerTEBCA() 
# # adicionarInformacion(dataFrameFacturas, dataFrameFacturasMAPFRE, dataFrameOncocenter, dataFrameTEBCA)
leerGrupoArticulo()